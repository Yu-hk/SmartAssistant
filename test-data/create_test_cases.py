from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

hfont = Font(bold=True, color='FFFFFF', size=11, name='Arial')
hfill = PatternFill('solid', fgColor='4472C4')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
calign = Alignment(vertical='center', wrap_text=True)
border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

p0f = PatternFill('solid', fgColor='FFC7CE')
p1f = PatternFill('solid', fgColor='FFEB9C')
p2f = PatternFill('solid', fgColor='C6EFCE')
tfills = {'正常流程': PatternFill('solid', fgColor='E2EFDA'), '异常场景': PatternFill('solid', fgColor='FCE4EC'), '边界场景': PatternFill('solid', fgColor='FFF3E0'), '权限场景': PatternFill('solid', fgColor='E8EAF6'), '安全场景': PatternFill('solid', fgColor='F3E5F5'), '兼容性场景': PatternFill('solid', fgColor='E0F7FA')}
pfills = {'P0': p0f, 'P1': p1f, 'P2': p2f}

def ss(ws, headers, widths, data):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hfont, hfill, halign, border
        ws.column_dimensions[get_column_letter(c)].width = w
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment, cell.border = calign, border
            if c == len(row)-1 and val in pfills: cell.fill = pfills[val]
            if c == len(row) and val in tfills: cell.fill = tfills[val]
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

h = ['用例ID', '模块', '功能点', '测试场景', '前置条件', '测试步骤', '预期结果', '优先级', '类型']
w = [12, 10, 14, 20, 20, 40, 35, 8, 12]

# Sheet 1: 概览
ws1 = wb.active
ws1.title = '测试计划概览'
ss(ws1, ['模块', '服务名', '端口', '核心功能', '测试用例数', '负责人'], [12, 28, 10, 50, 10, 12], [
    ['Gateway','smart-assistant-gateway','8081','JWT认证、路由转发、限流、Token管理',15,'待分配'],
    ['Consumer','smart-assistant-consumer','8082','会话管理、对话API、SSE流式、缓存、数据查询、分析面板',30,'待分配'],
    ['Router','smart-assistant-router','8083','意图识别、多Agent路由、语义缓存、任务规划、反思器',25,'待分配'],
    ['Order','smart-assistant-order','8085','订单查询、退款(二阶段确认)、物流追踪、旅游景点RAG',20,'待分配'],
    ['Product','smart-assistant-product','8084','商品查询、库存查询、价格查询、美食推荐RAG',18,'待分配'],
    ['User','smart-assistant-user','8086','用户注册、登录、JWT签发、会话管理',15,'待分配'],
    ['General','smart-assistant-general','8087','天气/新闻/搜索/计算/单位转换/汇率/文生图/图片解读',28,'待分配'],
    ['Common','smart-assistant-common','共享库','ReAct Agent/BGE嵌入/SQL安全校验/中文分词/@Tool日志切面',12,'待分配'],
    ['Frontend','React+TypeScript','3001','聊天界面、SSE流式展示、WebSocket、管理后台',20,'待分配'],
])

# Gateway
ss(wb.create_sheet('Gateway测试用例'), h, w, [
    ['GW-TC-001','Gateway','JWT认证','正常Token认证','已注册用户','1.调用/login获取Token 2.带Bearer Token调用受保护接口','返回200，正常处理','P0','正常流程'],
    ['GW-TC-002','Gateway','JWT认证','无Token请求','未提供Token','直接调用受保护接口（不带Authorization Header）','返回401 Unauthorized','P0','异常场景'],
    ['GW-TC-003','Gateway','JWT认证','过期Token','Token已过期','使用过期Token调用接口','返回401 Token过期','P1','异常场景'],
    ['GW-TC-004','Gateway','JWT认证','伪造Token','签名无效','使用伪造的JWT Token调用接口','返回401 签名无效','P1','异常场景'],
    ['GW-TC-005','Gateway','JWT认证','空Token','Token为空','设置Authorization: Bearer 空字符串','返回401','P1','边界场景'],
    ['GW-TC-006','Gateway','路由转发','服务不可用','目标服务Down','停止目标服务后发送请求到Gateway','返回502或错误提示','P1','异常场景'],
    ['GW-TC-007','Gateway','限流','正常限流','请求频率超限','短时间内发送大量请求','触发限流，返回429','P1','异常场景'],
    ['GW-TC-008','Gateway','Token管理','刷新Token','Token即将过期','调用刷新Token接口','返回新的有效Token','P1','正常流程'],
    ['GW-TC-009','Gateway','健康检查','健康状态','服务运行','GET /actuator/health','返回UP状态','P2','正常流程'],
    ['GW-TC-010','Gateway','跨域','CORS请求','前端跨域','从不同Origin发送请求','返回正确CORS Header','P2','兼容性场景'],
    ['GW-TC-011','Gateway','路由转发','请求体超限','超大请求体','POST超大Body','返回413','P2','边界场景'],
])

# Consumer
ss(wb.create_sheet('Consumer测试用例'), h, w, [
    ['CON-TC-001','Consumer','对话API','正常对话(无Session)','服务运行','POST /api/math/chat {"message":"你好"}','返回reply/suggestions','P0','正常流程'],
    ['CON-TC-002','Consumer','对话API','带Session对话','已有Session','POST /api/math/chat 带sessionId','返回reply/suggestions/sessionId','P0','正常流程'],
    ['CON-TC-003','Consumer','对话API','空消息','空消息','POST /api/math/chat {"message":""}','返回错误提示','P1','异常场景'],
    ['CON-TC-004','Consumer','对话API','超长消息','超过10万字符','POST超长消息','正确处理不崩溃','P2','边界场景'],
    ['CON-TC-005','Consumer','对话API','匿名用户','未认证','不传X-User-Id Header','使用anonymous处理','P1','权限场景'],
    ['CON-TC-006','Consumer','SSE流式','正常流式对话','服务运行','GET /api/math/stream/chat?message=你好','收到waiting→processing→response→done','P0','正常流程'],
    ['CON-TC-007','Consumer','SSE流式','请求排队','并发>5','同时发送6个流式请求','第6个收到queued事件','P1','异常场景'],
    ['CON-TC-008','Consumer','缓存','答案缓存命中','相同问题','第一次发送问题A→第二次相同A','第二次fromCache=true','P1','正常流程'],
    ['CON-TC-009','Consumer','缓存','缓存未命中','新问题','发送新问题B','fromCache=false','P1','正常流程'],
    ['CON-TC-010','Consumer','缓存','缓存统计(管理)','管理员','POST /api/math/cache/stats (ADMIN)','返回统计','P1','权限场景'],
    ['CON-TC-011','Consumer','缓存','缓存统计(用户)','普通用户','POST /api/math/cache/stats (USER)','返回403','P1','权限场景'],
    ['CON-TC-012','Consumer','数据查询','管理员查询','管理员','POST /api/data/query {"message":"查用户总数"}','返回查询结果','P1','正常流程'],
    ['CON-TC-013','Consumer','数据查询','非管理员查询','普通用户','POST /api/data/query (USER)','返回403','P1','权限场景'],
    ['CON-TC-014','Consumer','数据查询','SQL注入','注入语句','POST /api/data/query 含SQL注入','SQL安全校验器拦截','P0','安全场景'],
    ['CON-TC-015','Consumer','限流','API限流','超过30req/s','1秒内发31个请求','第31个被限流','P1','异常场景'],
    ['CON-TC-016','Consumer','分析面板','获取统计','有数据','GET /api/analytics/stats','返回统计','P2','正常流程'],
    ['CON-TC-017','Consumer','分析面板','热门建议','有点击','GET /api/analytics/top-suggestions','返回前5','P2','正常流程'],
    ['CON-TC-018','Consumer','分析面板','偏好分布','有记录','GET /api/analytics/preference-distribution','返回比例','P2','正常流程'],
    ['CON-TC-019','Consumer','对话价值','有价值对话','>=3轮+工具','多轮对话触发工具调用','生成个人文档','P2','正常流程'],
    ['CON-TC-020','Consumer','对话价值','无价值对话','单轮','简单问答无工具调用','不生成文档','P2','正常流程'],
])

# Router
ss(wb.create_sheet('Router测试用例'), h, w, [
    ['ROU-TC-001','Router','路由决策','订单意图','Router运行','POST /api/router/route {"question":"查订单ORD-2024001"}','agentName=order','P0','正常流程'],
    ['ROU-TC-002','Router','路由决策','商品意图','Router运行','POST {"question":"iPhone 15 Pro多少钱？"}','agentName=product','P0','正常流程'],
    ['ROU-TC-003','Router','路由决策','天气意图','Router运行','POST {"question":"北京今天天气？"}','agentName=general','P0','正常流程'],
    ['ROU-TC-004','Router','路由决策','多Agent协作','Router运行','POST {"question":"推荐北京景点和川菜"}','触发多Agent协作','P0','正常流程'],
    ['ROU-TC-005','Router','路由决策','模糊意图','模糊问题','POST {"question":"嗯"}','合理兜底','P2','异常场景'],
    ['ROU-TC-006','Router','路由决策','空问题','空消息','POST {"question":""}','参数校验错误','P1','异常场景'],
    ['ROU-TC-007','Router','反思器','五维评分','低质量回复','模拟低质量Agent回复','评分<0.6切换fallback','P1','正常流程'],
    ['ROU-TC-008','Router','反思器','高质量回复','高质量回复','模拟高质量Agent回复','评分>=0.6直接返回','P1','正常流程'],
    ['ROU-TC-009','Router','语义缓存','缓存命中','相同问题','第一次发A→第二次相同A','fromCache=true','P1','正常流程'],
    ['ROU-TC-010','Router','语义缓存','短TTL跳过','TTL<1h','短时效Agent回复','跳过缓存','P2','正常流程'],
    ['ROU-TC-011','Router','语义缓存','缓存未命中','不同问题','发送不同问题','fromCache=false','P1','正常流程'],
    ['ROU-TC-012','Router','RAG增强','旅游知识','启用RAG','发旅游问题','触发RAG检索','P1','正常流程'],
    ['ROU-TC-013','Router','任务规划','多意图分解','复杂问题','POST多意图问题','分解为子任务','P0','正常流程'],
    ['ROU-TC-014','Router','任务规划','并行执行','多子任务','执行子任务','最大4并发','P1','正常流程'],
    ['ROU-TC-015','Router','任务规划','结果合并','多Agent返回','多个子任务完成','ResultMerger整合','P0','正常流程'],
    ['ROU-TC-016','Router','任务规划','部分失败','Agent不可用','某个Agent失败','成功部分返回','P1','异常场景'],
    ['ROU-TC-017','Router','Agent发现','动态发现','新Agent注册','通过Nacos注册','自动发现缓存','P1','正常流程'],
    ['ROU-TC-018','Router','Agent发现','健康检查','不健康','检测到不健康Agent','标记不可用','P1','异常场景'],
    ['ROU-TC-019','Router','健康检查','状态','运行中','GET /api/router/health','返回running','P2','正常流程'],
])

# Order
ss(wb.create_sheet('Order测试用例'), h, w, [
    ['ORD-TC-001','Order','查询订单','正常查询','有数据','queryOrder("ORD-2024001")','完整订单信息','P0','正常流程'],
    ['ORD-TC-002','Order','查询订单','不存在订单','无效单号','queryOrder("ORD-NOT-EXIST")','ORDER_NOT_FOUND','P0','异常场景'],
    ['ORD-TC-003','Order','查询订单','空订单号','空','queryOrder("")','订单不存在','P1','边界场景'],
    ['ORD-TC-004','Order','查询订单','各种状态','遍历','查已发货/待发货/已签收/退款中','正确显示','P0','正常流程'],
    ['ORD-TC-005','Order','退款申请','正常流程','已签收','applyRefund("ORD-2024003","质量问题")','需确认提醒','P0','正常流程'],
    ['ORD-TC-006','Order','退款申请','二阶段确认','已确认','confirmAction→重新applyRefund','退款成功','P0','正常流程'],
    ['ORD-TC-007','Order','退款申请','跳过确认','未确认','不调confirmAction直接退款','被阻止','P0','异常场景'],
    ['ORD-TC-008','Order','物流追踪','正常查询','有运单号','trackLogistics("SF1234567890")','物流轨迹','P0','正常流程'],
    ['ORD-TC-009','Order','物流追踪','空运单号','空','trackLogistics("")','TRACKING_REQUIRED','P1','异常场景'],
    ['ORD-TC-010','Order','RAG','旅游景点','启用RAG','查"北京故宫"','向量检索','P1','正常流程'],
    ['ORD-TC-011','Order','多轮对话','多轮查询','连续对话','先查A再追问物流','保持上下文','P0','正常流程'],
])

# Product
ss(wb.create_sheet('Product测试用例'), h, w, [
    ['PRO-TC-001','Product','商品查询','正常查询','有数据','queryProductInfo("IPHONE-15-PRO")','商品详细信息','P0','正常流程'],
    ['PRO-TC-002','Product','商品查询','不存在商品','无效编码','queryProductInfo("NONEXIST")','PRODUCT_NOT_FOUND','P0','异常场景'],
    ['PRO-TC-003','Product','商品查询','模糊匹配','部分匹配','queryProductInfo("iPhone")','自动匹配','P1','正常流程'],
    ['PRO-TC-004','Product','商品查询','空编码','空','queryProductInfo("")','未找到','P1','边界场景'],
    ['PRO-TC-005','Product','库存查询','库存充足','充足','checkStock("IPHONE-15-PRO")','24小时内发货','P0','正常流程'],
    ['PRO-TC-006','Product','库存查询','库存紧张','紧张','checkStock("MACBOOK-AIR-M3")','建议尽快下单','P0','正常流程'],
    ['PRO-TC-007','Product','价格查询','正常查询','有价格','getPrice("IPHONE-15-PRO")','售价+分期','P0','正常流程'],
    ['PRO-TC-008','Product','多轮对话','商品推荐','有偏好','"推荐手机"→追问→推荐','推荐相关','P0','正常流程'],
    ['PRO-TC-009','Product','知识库','特色菜系','有数据','查"北京特色菜"','知识库检索','P1','正常流程'],
    ['PRO-TC-010','Product','RAG','美食评论','有评论','查"北京烤鸭推荐"','搜索评论','P1','正常流程'],
])

# User
ss(wb.create_sheet('User测试用例'), h, w, [
    ['USR-TC-001','User','注册','正常注册','新用户','POST /api/auth/register 合法信息','返回Token','P0','正常流程'],
    ['USR-TC-002','User','注册','重复用户名','已存在','已存在用户名注册','返回400','P1','异常场景'],
    ['USR-TC-003','User','注册','弱密码','密码不足','密码"123"','密码强度不通过','P1','异常场景'],
    ['USR-TC-004','User','登录','正常登录','已注册','POST /api/auth/login 正确凭据','返回Token','P0','正常流程'],
    ['USR-TC-005','User','登录','错误密码','密码错','POST /api/auth/login 错误密码','返回401','P0','异常场景'],
    ['USR-TC-006','User','登录','不存在用户','未注册','POST /api/auth/login 未注册用户','返回401','P1','异常场景'],
    ['USR-TC-007','User','获取用户','有效Token','已登录','GET /api/auth/me 带Token','返回用户信息','P0','正常流程'],
    ['USR-TC-008','User','获取用户','过期Token','已过期','GET /api/auth/me 过期Token','返回401','P1','异常场景'],
    ['USR-TC-009','User','JWT','签名验证','篡改','使用不同签名Token','签名无效返回401','P0','异常场景'],
])

# General
ss(wb.create_sheet('General测试用例'), h, w, [
    ['GEN-TC-001','General','数学计算','基础运算','—','calculate("3.14*5^2")','78.5','P0','正常流程'],
    ['GEN-TC-002','General','数学计算','平方根','—','calculate("sqrt(144)")','12','P0','正常流程'],
    ['GEN-TC-003','General','数学计算','复合运算','—','calculate("(12+8)*3.5/2")','35','P0','正常流程'],
    ['GEN-TC-004','General','数学计算','除零','—','calculate("1/0")','无穷大/错误','P1','异常场景'],
    ['GEN-TC-005','General','数学计算','无效式','—','calculate("abc")','解析错误','P1','异常场景'],
    ['GEN-TC-006','General','温度转换','C→F','—','convertTemperature(100,C,F)','212F','P0','正常流程'],
    ['GEN-TC-007','General','温度转换','F→C','—','convertTemperature(32,F,C)','0C','P0','正常流程'],
    ['GEN-TC-008','General','温度转换','C→K','—','convertTemperature(0,C,K)','273.15K','P1','正常流程'],
    ['GEN-TC-009','General','长度转换','km→m','—','convertLength(1,km,m)','1000 m','P0','正常流程'],
    ['GEN-TC-010','General','重量转换','kg→lb','—','convertWeight(1,kg,lb)','2.2046 lb','P1','正常流程'],
    ['GEN-TC-011','General','天气查询','正常城市','网络可达','queryWeather("北京")','温度/湿度/风力','P0','正常流程'],
    ['GEN-TC-012','General','天气查询','英文城市','网络可达','queryWeather("London")','伦敦天气','P1','正常流程'],
    ['GEN-TC-013','General','天气查询','无效城市','错误','queryWeather("XYZ123")','WEATHER_UNAVAILABLE','P1','异常场景'],
    ['GEN-TC-014','General','新闻热点','获取热点','网络可达','getHotNews()','热点榜单','P0','正常流程'],
    ['GEN-TC-015','General','联网搜索','正常搜索','网络可达','searchWeb("2025年春节")','搜索结果','P0','正常流程'],
    ['GEN-TC-016','General','汇率转换','USD→CNY','网络可达','convertCurrency(100,USD,CNY)','人民币+汇率','P0','正常流程'],
    ['GEN-TC-017','General','汇率转换','相同货币','—','convertCurrency(100,CNY,CNY)','100 CNY','P1','边界场景'],
    ['GEN-TC-018','General','汇率转换','无效货币','—','convertCurrency(100,INVALID,CNY)','INVALID_CURRENCY','P1','异常场景'],
    ['GEN-TC-019','General','历史纠错','有记录','有历史','queryCorrections("世界最高峰")','返回修正','P1','正常流程'],
    ['GEN-TC-020','General','计算脚本','正常脚本','—','executeScript("a=3\\nb=4\\nc=sqrt(a^2+b^2)")','c=5','P1','正常流程'],
    ['GEN-TC-021','General','计算脚本','注入防护','恶意','含import/exec关键词','SCRIPT_REJECTED','P0','异常场景'],
])

# Frontend
ss(wb.create_sheet('前端测试用例'), h, w, [
    ['FE-TC-001','前端','聊天界面','发送消息','已登录','输入消息→发送','显示回复','P0','正常流程'],
    ['FE-TC-002','前端','聊天界面','SSE流式','流式回复','发送消息','逐步显示thinking/response','P0','正常流程'],
    ['FE-TC-003','前端','聊天界面','空消息','空输入','不输入直接发送','按钮禁用或提示','P1','异常场景'],
    ['FE-TC-004','前端','聊天界面','排队提示','繁忙','排队时发送','显示位置和等待时间','P1','正常流程'],
    ['FE-TC-005','前端','聊天界面','转人工','需人工','点转人工','连接客服','P1','正常流程'],
    ['FE-TC-006','前端','聊天界面','满意度','结束','评价→选分','提交成功','P2','正常流程'],
    ['FE-TC-007','前端','会话管理','新建','—','点新建会话','清空聊天记录','P0','正常流程'],
    ['FE-TC-008','前端','会话管理','切换','多会话','点历史会话','切换到该会话','P0','正常流程'],
    ['FE-TC-009','前端','会话管理','删除','有记录','删除会话','永久删除','P1','正常流程'],
    ['FE-TC-010','前端','管理后台','查看统计','管理员','进入管理后台','统计面板','P1','正常流程'],
    ['FE-TC-011','前端','登录注册','登录','—','正确凭据','跳转聊天页','P0','正常流程'],
    ['FE-TC-012','前端','登录注册','注册','—','填信息提交','注册成功自动登录','P0','正常流程'],
    ['FE-TC-013','前端','登录注册','登录失败','错误','错误密码','错误提示','P1','异常场景'],
])

# Security
ss(wb.create_sheet('安全兼容性测试'), h, w, [
    ['SEC-TC-001','安全','SQL注入','注入攻击','—','数据查询接口发SQL注入','jsqlparser拦截','P0','安全场景'],
    ['SEC-TC-002','安全','SQL注入','表名白名单','—','查不在白名单的表','Validator拦截','P0','安全场景'],
    ['SEC-TC-003','安全','工具权限','AdminOnly','非管理员','非管理员调@AdminOnly','切面拦截','P1','权限场景'],
    ['SEC-TC-004','安全','工具权限','二阶段确认','敏感操作','退款未确认前','ApprovalService阻止','P0','安全场景'],
    ['SEC-TC-005','安全','JWT','Token伪造','篡改','改Token payload','签名校验失败','P0','安全场景'],
    ['SEC-TC-006','兼容性','浏览器','Chrome','—','Chrome打开前端','正常','P2','兼容性场景'],
    ['SEC-TC-007','兼容性','浏览器','Firefox','—','Firefox打开前端','正常','P2','兼容性场景'],
    ['SEC-TC-008','兼容性','屏幕','窄屏1024','—','缩窄到1024px','布局正常','P2','兼容性场景'],
])

wb.save('/workspace/SmartAssistant/test-data/SmartAssistant_测试用例表.xlsx')
print('OK')
